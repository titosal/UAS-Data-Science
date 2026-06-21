import pandas as pd
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load the original data
df = pd.read_csv('ulasan_gacoan_tanggal_jak_boul.csv')

# Clean the text
def clean_text(text):
    if not isinstance(text, str):
        return ""
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.lower().strip()

df['cleaned_review'] = df['review'].apply(clean_text)
df.dropna(subset=['cleaned_review'], inplace=True)

# Separate the data
df_positif = df[df['rating'] >= 4].copy()
df_negatif = df[df['rating'] < 4].copy()

# Generate Excel File
file_name = "Hasil_Analisis_Ulasan_Gacoan_tanggal.xlsx"

with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    # Write to different sheets
    df.to_excel(writer, sheet_name='Semua Data', index=False)
    df_positif.to_excel(writer, sheet_name='Ulasan Positif', index=False)
    df_negatif.to_excel(writer, sheet_name='Ulasan Negatif', index=False)
    
    workbook = writer.book
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    # Using a professional muted blue-grey for headers
    header_fill = PatternFill(start_color="3B4252", end_color="3B4252", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='D8DEE9'),
                         right=Side(style='thin', color='D8DEE9'),
                         top=Side(style='thin', color='D8DEE9'),
                         bottom=Side(style='thin', color='D8DEE9'))
    
    wrap_alignment = Alignment(vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='top')

    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        # Freeze top row
        worksheet.freeze_panes = 'A2'
        
        # Apply AutoFilter
        worksheet.auto_filter.ref = worksheet.dimensions

        # Style headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
            
        # Define column widths based on expected content
        col_widths = {'A': 15, 'B': 25, 'C': 60, 'D': 10, 'E': 60}
        
        for col_letter, width in col_widths.items():
            worksheet.column_dimensions[col_letter].width = width
            
        # Apply styles to data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                if cell.column_letter == 'D': # Rating column
                    cell.alignment = center_alignment
                else:
                    cell.alignment = wrap_alignment

print(f"Excel file created: {file_name}")